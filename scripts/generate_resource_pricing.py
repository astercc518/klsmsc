#!/usr/bin/env python3
"""
根据 成本 (2).xlsx 生成资源报价信息

Excel 结构: 类型, 供应商, 国家, 成本价（U）, 销售价（U）, 描述
输出: data/resource_pricing.json（按供应商、国家组织的报价数据）

用法:
  docker compose run --rm -v $(pwd):/workspace -w /workspace api bash -c \
    "pip install openpyxl -q && python scripts/generate_resource_pricing.py"
"""
import json
import os
import sys
from collections import defaultdict
from datetime import date

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_SCRIPT_DIR)

# 复用 import_supplier_pricing 的国家映射与解析逻辑
sys.path.insert(0, os.path.join(_ROOT, 'scripts'))
from import_supplier_pricing import (  # noqa: E402
    COUNTRY_CN_TO_CODE,
    parse_price,
    supplier_to_channel_code,
)


def main():
    try:
        import openpyxl
    except ImportError:
        print("请先安装: pip install openpyxl")
        sys.exit(1)

    xlsx_path = os.path.join(_ROOT, '成本 (2).xlsx')
    if not os.path.exists(xlsx_path):
        print(f"文件不存在: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if '价格' not in wb.sheetnames:
        print("Excel 中未找到「价格」工作表")
        sys.exit(1)

    ws = wb['价格']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows or rows[0][:3] != ('类型', '供应商', '国家'):
        print("表头格式不符，期望: 类型, 供应商, 国家, 成本价（U）, 销售价（U）, 描述")
        sys.exit(1)

    # 解析所有类型（SMS 及其他）
    by_supplier = defaultdict(list)
    by_country = defaultdict(list)
    all_items = []
    skipped_countries = set()

    for r in rows[1:]:
        if not r or len(r) < 5:
            continue
        typ = (r[0] or '').strip()
        supplier = (r[1] or '').strip()
        country_cn = (r[2] or '').strip()
        cost = parse_price(r[3])  # 成本价
        sale = parse_price(r[4])  # 销售价
        desc = (r[5] or '').strip() if len(r) > 5 else ''

        if not supplier or not country_cn:
            continue
        # 销售价优先，无则用成本价
        price = sale if sale is not None and sale > 0 else cost
        if price is None or price <= 0:
            continue

        country_code = COUNTRY_CN_TO_CODE.get(country_cn)
        if not country_code:
            skipped_countries.add(country_cn)
            continue

        channel_code = supplier_to_channel_code(supplier)
        item = {
            'type': typ,
            'supplier': supplier,
            'channel_code': channel_code,
            'country': country_cn,
            'country_code': country_code,
            'cost_usd': round(cost, 4) if cost is not None else None,
            'sale_usd': round(sale, 4) if sale is not None else None,
            'price_usd': round(price, 4),
            'description': desc or None,
        }
        all_items.append(item)
        by_supplier[supplier].append(item)
        by_country[country_code].append(item)

    if skipped_countries:
        print(f"跳过未映射国家: {', '.join(sorted(skipped_countries))}")

    # 按供应商汇总（每个国家取最低价）
    supplier_summary = {}
    for supplier, items in by_supplier.items():
        countries = {}
        for it in items:
            cc = it['country_code']
            if cc not in countries or it['price_usd'] < countries[cc]['price_usd']:
                countries[cc] = it
        supplier_summary[supplier] = {
            'channel_code': supplier_to_channel_code(supplier),
            'country_count': len(countries),
            'countries': {cc: v for cc, v in sorted(countries.items())},
        }

    # 按国家汇总（每个国家各供应商报价，按价格排序）
    country_summary = {}
    for cc, items in by_country.items():
        sorted_items = sorted(items, key=lambda x: x['price_usd'])
        country_summary[cc] = {
            'country_name': sorted_items[0]['country'] if sorted_items else '',
            'supplier_count': len(set(it['supplier'] for it in items)),
            'lowest_price': round(sorted_items[0]['price_usd'], 4) if sorted_items else None,
            'offers': [
                {
                    'supplier': it['supplier'],
                    'channel_code': it['channel_code'],
                    'price_usd': it['price_usd'],
                    'cost_usd': it['cost_usd'],
                    'sale_usd': it['sale_usd'],
                }
                for it in sorted_items[:10]  # 每国家最多 10 条
            ],
        }

    output = {
        'generated_at': date.today().isoformat(),
        'source': '成本 (2).xlsx',
        'total_records': len(all_items),
        'supplier_count': len(by_supplier),
        'country_count': len(by_country),
        'currency': 'USD',
        'by_supplier': dict(supplier_summary),
        'by_country': dict(country_summary),
        'flat_list': all_items,
    }

    out_dir = os.path.join(_ROOT, 'data')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'resource_pricing.json')

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"已生成资源报价信息: {out_path}")
    print(f"  总记录: {len(all_items)}, 供应商: {len(by_supplier)}, 国家: {len(by_country)}")


if __name__ == '__main__':
    main()

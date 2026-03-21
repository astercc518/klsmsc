#!/usr/bin/env python3
"""
从 成本 (2).xlsx 导入供应商报价到 country_pricing 表
供业务助手 Bot 报价查询使用

Excel 结构: 类型, 供应商, 国家, 成本价（U）, 销售价（U）, 描述
仅导入类型=SMS 的记录，供应商作为通道名称展示

用法: 在项目根目录执行
  docker compose run --rm -v $(pwd):/workspace -w /workspace api bash -c \
    "pip install openpyxl -q && python scripts/import_supplier_pricing.py"
"""
import os
import sys
from decimal import Decimal
from datetime import date

# 添加 backend 到 path
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_SCRIPT_DIR)
_BACKEND = os.path.join(_ROOT, 'backend')
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)

# 国家中文名 -> ISO 3166 二位代码
COUNTRY_CN_TO_CODE = {
    '中国': 'CN', '美国': 'US', '英国': 'GB', '新加坡': 'SG', '日本': 'JP', '韩国': 'KR',
    '泰国': 'TH', '越南': 'VN', '马来西亚': 'MY', '印尼': 'ID', '印度尼西亚': 'ID',
    '菲律宾': 'PH', '印度': 'IN', '澳大利亚': 'AU', '加拿大': 'CA', '德国': 'DE',
    '法国': 'FR', '意大利': 'IT', '西班牙': 'ES', '俄罗斯': 'RU', '巴西': 'BR',
    '墨西哥': 'MX', '香港': 'HK', '台湾': 'TW', '阿联酋': 'AE', '阿拉伯联合酋长国': 'AE',
    '沙特': 'SA', '土耳其': 'TR', '哥伦比亚': 'CO', '秘鲁': 'PE', '尼日利亚': 'NG',
    '孟加拉': 'BD', '巴基斯坦': 'PK', '埃及': 'EG', '南非': 'ZA', '肯尼亚': 'KE',
    '加纳': 'GH', '摩洛哥': 'MA', '阿尔及利亚': 'DZ', '突尼斯': 'TN', '乌干达': 'UG',
    '坦桑尼亚': 'TZ', '埃塞俄比亚': 'ET', '埃塞尔比亚': 'ET', '塞内加尔': 'SN',
    '喀麦隆': 'CM', '科特迪瓦': 'CI', '赞比亚': 'ZM', '以色列': 'IL', '约旦': 'JO',
    '卡塔尔': 'QA', '科威特': 'KW', '阿曼': 'OM', '巴林': 'BH', '黎巴嫩': 'LB',
    '丹麦': 'DK', '瑞典': 'SE', '挪威': 'NO', '芬兰': 'FI', '荷兰': 'NL',
    '比利时': 'BE', '瑞士': 'CH', '奥地利': 'AT', '波兰': 'PL', '捷克': 'CZ',
    '罗马尼亚': 'RO', '匈牙利': 'HU', '希腊': 'GR', '葡萄牙': 'PT', '爱尔兰': 'IE',
    '新西兰': 'NZ', '阿根廷': 'AR', '智利': 'CL', '委内瑞拉': 'VE', '厄瓜多尔': 'EC',
    '玻利维亚': 'BO', '柬埔寨': 'KH', '老挝': 'LA', '缅甸': 'MM', '缅甸 ': 'MM',
    '文莱': 'BN', '斯里兰卡': 'LK', '哈萨克斯坦': 'KZ', '立陶宛': 'LT',
    '美国+1': 'US',
}


def parse_price(val):
    """解析价格"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('\u202c', '').replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def supplier_to_channel_code(name: str) -> str:
    """供应商名转唯一通道编码"""
    mapping = {
        '一正通信': 'SUP_YIZHENG', 'TS通信': 'SUP_TS', 'NU通信': 'SUP_NU',
        'KMI通信': 'SUP_KMI', '黑豹通信': 'SUP_HEIBAO', '节点通信': 'SUP_JIEDIAN',
        'QK通信': 'SUP_QK', 'DL通信': 'SUP_DL', '创优通信': 'SUP_CHUANGYOU',
        'BO通信': 'SUP_BO', '左总通信': 'SUP_ZUOZONG', '速运通信': 'SUP_SUYUN',
        'AIY通信': 'SUP_AIY', '葵芳通信': 'SUP_KUIFANG', 'ANY通信': 'SUP_ANY',
        'WOID通信': 'SUP_WOID', '众诚通信': 'SUP_ZHONGCHENG', 'BE通信': 'SUP_BE',
        'SKY通信': 'SUP_SKY', '粤讯通信': 'SUP_YUEXUN', '玄武通信': 'SUP_XUANWU',
        'ueasy通信': 'SUP_UEASY', '九天通信': 'SUP_JIUTIAN',
    }
    return mapping.get(name, 'SUP_' + name[:10].replace(' ', '_').replace('通信', ''))


def main():
    try:
        import openpyxl
    except ImportError:
        print("请先安装: pip install openpyxl")
        sys.exit(1)

    try:
        import pymysql
    except ImportError:
        print("请确保在 api 容器中运行（含 pymysql）")
        sys.exit(1)

    xlsx_path = os.path.join(_ROOT, '成本 (2).xlsx')
    if not os.path.exists(xlsx_path):
        print(f"文件不存在: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if '价格' not in wb.sheetnames:
        print("Excel 中未找到「价格」工作表")
        sys.exit(1)

    ws = wb['价格']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows or rows[0][:3] != ('类型', '供应商', '国家'):
        print("表头格式不符，期望: 类型, 供应商, 国家, 成本价（U）, 销售价（U）, 描述")
        sys.exit(1)

    # 仅处理 SMS
    sms_rows = []
    for r in rows[1:]:
        if not r or len(r) < 5:
            continue
        typ = (r[0] or '').strip()
        supplier = (r[1] or '').strip()
        country = (r[2] or '').strip()
        price = parse_price(r[4])
        if typ != 'SMS' or not supplier or not country or price is None or price <= 0:
            continue
        cc = COUNTRY_CN_TO_CODE.get(country)
        if not cc:
            print(f"  跳过未映射国家: {country}")
            continue
        sms_rows.append((supplier, country, cc, round(price, 4)))

    print(f"共解析到 {len(sms_rows)} 条 SMS 报价")

    host = os.environ.get('DATABASE_HOST', 'localhost')
    port = int(os.environ.get('DATABASE_PORT', 3306))
    user = os.environ.get('DATABASE_USER', 'smsuser')
    pwd = os.environ.get('DATABASE_PASSWORD', 'smspass123')
    db = os.environ.get('DATABASE_NAME', 'sms_system')

    conn = pymysql.connect(host=host, port=port, user=user, password=pwd, db=db, charset='utf8mb4')

    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            # 1. 为每个供应商创建通道
            supplier_channels = {}
            today = date.today().isoformat()

            for supplier, country_cn, country_code, price in sms_rows:
                if supplier in supplier_channels:
                    continue
                code = supplier_to_channel_code(supplier)
                cur.execute(
                    "SELECT id FROM channels WHERE channel_code = %s",
                    (code,)
                )
                row = cur.fetchone()
                if row:
                    supplier_channels[supplier] = row['id']
                else:
                    cur.execute(
                        """INSERT INTO channels (channel_code, channel_name, protocol, default_sender_id, status)
                           VALUES (%s, %s, 'HTTP', 'QUOTE', 'inactive')""",
                        (code, supplier)
                    )
                    supplier_channels[supplier] = cur.lastrowid
                    print(f"  新建通道: {code} ({supplier})")

            conn.commit()

            # 2. 插入 country_pricing（同通道+国家仅保留一条，先删后插）
            seen = set()
            inserted = 0
            for supplier, country_cn, country_code, price in sms_rows:
                ch_id = supplier_channels[supplier]
                key = (ch_id, country_code)
                if key in seen:
                    continue
                seen.add(key)

                cur.execute(
                    "DELETE FROM country_pricing WHERE channel_id = %s AND country_code = %s",
                    (ch_id, country_code)
                )
                cur.execute(
                    """INSERT INTO country_pricing (channel_id, country_code, country_name, price_per_sms, currency, effective_date)
                       VALUES (%s, %s, %s, %s, 'USD', %s)""",
                    (ch_id, country_code, country_cn, str(price), today)
                )
                inserted += 1

            conn.commit()
            print(f"新增 {inserted} 条报价记录")

    finally:
        conn.close()

    print("导入完成，销售可在 Bot 报价查询中查看")


if __name__ == '__main__':
    main()
